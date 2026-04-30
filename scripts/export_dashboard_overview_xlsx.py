"""
One-off export: PromptWars in-person and virtual stats to .xlsx.

Sheets (in order):
  1. Summary               High-level numbers for both tracks.
  2. InPerson              Detailed in-person stats + in-person -> virtual overlap.
  3. InPerson_Cities       City-wise registrations (one row per city).
  4. InPerson_Top10        Top 10 teams (main challenge) with deployed links; top 3 highlighted.
  5. Virtual               Detailed virtual stats incl. cross-challenge overlap metrics.
  6. Virtual_Challenges    Per-challenge: registrations at open / close, totals & unique submissions.
  7. V_C{id}_<slug>        One sheet per virtual challenge with top-N participants; top 10 highlighted.

Cross-challenge metrics use distinct ``leader_email_normalized`` in
``virtual_challenge_submission_rows``. In-person <-> Virtual overlap matches
``email_normalized`` on the two main registration tables.

The "first" and "second" virtual challenges (used for cross-challenge metrics)
are auto-detected by ``opens_at ASC NULLS LAST, id ASC`` for the given virtual
event. Use ``--virtual-c1-id`` / ``--virtual-c2-id`` to override.

Usage (from repo root, with ``.env`` / ``DATABASE_URL`` set)::

    python scripts/export_dashboard_overview_xlsx.py
    python scripts/export_dashboard_overview_xlsx.py -o reports/snapshot.xlsx
    python scripts/export_dashboard_overview_xlsx.py --in-person-event-id 1 --virtual-event-id 2
    python scripts/export_dashboard_overview_xlsx.py --top-n 400 --virtual-c1-top-n 400
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv  # noqa: E402

load_dotenv(ROOT / ".env")

import app as pw  # noqa: E402  loads engine + table-name constants

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from sqlalchemy import text  # noqa: E402


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SECTION_FONT = Font(bold=True, color="0F172A")
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
TOP3_FONT = Font(bold=True, color="78350F")
TOP3_FILL = PatternFill("solid", fgColor="FDE68A")
TOP10_FONT = Font(bold=True, color="0F172A")
TOP10_FILL = PatternFill("solid", fgColor="FEF3C7")


def _slug(text_in: str | None, *, max_len: int = 18) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(text_in or "")).strip("_")
    return (s[:max_len] or "untitled").lower()


def _safe_sheet(name: str) -> str:
    s = re.sub(r"[\[\]*?:/\\]", "_", name)
    return s[:31] or "Sheet"


def _autosize(ws, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 10
        for cell in ws[letter]:
            if cell.value is None:
                continue
            w = min(max_width, len(str(cell.value)) + 2)
            best = max(best, w)
        ws.column_dimensions[letter].width = best


def _kv_section(ws, title: str, rows: Iterable[tuple[str, Any]]) -> None:
    r = ws.max_row + (2 if ws.max_row > 1 or ws.cell(row=1, column=1).value else 0)
    if r < 1:
        r = 1
    if title:
        ws.cell(row=r, column=1, value=title).font = SECTION_FONT
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.cell(row=r, column=2).fill = SECTION_FILL
        r += 1
    for k, v in rows:
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        r += 1


def _table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    title: str | None = None,
    highlight_top3: bool = False,
    highlight_top10: bool = False,
) -> None:
    r = ws.max_row + (2 if ws.max_row > 1 or ws.cell(row=1, column=1).value else 0)
    if r < 1:
        r = 1
    if title:
        ws.cell(row=r, column=1, value=title).font = SECTION_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        r += 1
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    header_row = r
    for i, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=header_row + i, column=c, value=val)
            if highlight_top3 and i <= 3:
                cell.font = TOP3_FONT
                cell.fill = TOP3_FILL
            elif highlight_top10 and i <= 10:
                cell.font = TOP10_FONT
                cell.fill = TOP10_FILL
    if not rows:
        ws.cell(row=header_row + 1, column=1, value="(no rows)")


def _scalar(conn, sql: str, params: dict | None = None) -> Any:
    return conn.execute(text(sql), params or {}).scalar()


def _all(conn, sql: str, params: dict | None = None) -> list[dict]:
    rows = conn.execute(text(sql), params or {}).mappings().all()
    return [dict(r) for r in rows]


def _virtual_challenges(conn, virtual_event_id: int) -> list[dict]:
    """Challenges of the virtual event ordered by opens_at ASC NULLS LAST, id ASC."""
    return _all(
        conn,
        """
        SELECT id, title, status, opens_at, closes_at, created_at
        FROM challenges
        WHERE event_id = :eid
        ORDER BY opens_at ASC NULLS LAST, id ASC
        """,
        {"eid": int(virtual_event_id)},
    )


def _virtual_event_total_registrations(conn, virtual_event_id: int) -> int:
    return int(
        _scalar(
            conn,
            f"SELECT COUNT(*) FROM {pw.TABLE_VIRTUAL_MDC} WHERE event_id = :eid",
            {"eid": int(virtual_event_id)},
        )
        or 0
    )


def _virtual_per_challenge_stats(conn, virtual_event_id: int, challenges: list[dict]) -> list[dict]:
    """For each challenge: registrations at open / close, total + unique submissions."""
    out: list[dict] = []
    for ch in challenges:
        eid = int(virtual_event_id)
        cid = int(ch["id"])
        opens_at = ch.get("opens_at")
        closes_at = ch.get("closes_at")
        reg_at_open = None
        reg_at_close = None
        if opens_at is not None:
            reg_at_open = int(
                _scalar(
                    conn,
                    f"""
                    SELECT COUNT(*) FROM {pw.TABLE_VIRTUAL_MDC}
                    WHERE event_id = :eid AND form_timestamp IS NOT NULL
                      AND form_timestamp <= :ts
                    """,
                    {"eid": eid, "ts": opens_at},
                )
                or 0
            )
        if closes_at is not None:
            reg_at_close = int(
                _scalar(
                    conn,
                    f"""
                    SELECT COUNT(*) FROM {pw.TABLE_VIRTUAL_MDC}
                    WHERE event_id = :eid AND form_timestamp IS NOT NULL
                      AND form_timestamp <= :ts
                    """,
                    {"eid": eid, "ts": closes_at},
                )
                or 0
            )
        sub_row = conn.execute(
            text(
                """
                SELECT COUNT(*)::bigint AS total_subs,
                       COUNT(DISTINCT leader_email_normalized)::bigint AS unique_subs,
                       COUNT(*) FILTER (WHERE deployed_link IS NOT NULL
                                              AND btrim(deployed_link) <> '')::bigint AS deployed,
                       COUNT(*) FILTER (WHERE total_score IS NOT NULL)::bigint AS scored
                FROM virtual_challenge_submission_rows
                WHERE event_id = :eid AND challenge_id = :cid
                """
            ),
            {"eid": eid, "cid": cid},
        ).mappings().fetchone() or {}
        out.append(
            {
                "challenge_id": cid,
                "title": ch.get("title"),
                "status": ch.get("status"),
                "opens_at": opens_at,
                "closes_at": closes_at,
                "registrations_at_open": reg_at_open,
                "registrations_at_close": reg_at_close,
                "registrations_during_window": (
                    None
                    if reg_at_open is None or reg_at_close is None
                    else int(reg_at_close) - int(reg_at_open)
                ),
                "total_submissions": int(sub_row.get("total_subs") or 0),
                "unique_submissions_by_email": int(sub_row.get("unique_subs") or 0),
                "deployed_count": int(sub_row.get("deployed") or 0),
                "scored_count": int(sub_row.get("scored") or 0),
            }
        )
    return out


def _virtual_overall_submission_stats(conn, virtual_event_id: int) -> dict[str, int]:
    row = conn.execute(
        text(
            """
            SELECT COUNT(*)::bigint AS total_subs,
                   COUNT(DISTINCT leader_email_normalized)::bigint AS unique_subs,
                   COUNT(*) FILTER (WHERE deployed_link IS NOT NULL
                                          AND btrim(deployed_link) <> '')::bigint AS deployed,
                   COUNT(*) FILTER (WHERE total_score IS NOT NULL)::bigint AS scored
            FROM virtual_challenge_submission_rows
            WHERE event_id = :eid
            """
        ),
        {"eid": int(virtual_event_id)},
    ).mappings().fetchone() or {}
    return {
        "total_submissions": int(row.get("total_subs") or 0),
        "unique_submissions_by_email": int(row.get("unique_subs") or 0),
        "deployed_count": int(row.get("deployed") or 0),
        "scored_count": int(row.get("scored") or 0),
    }


def _virtual_top_n_for_challenge(
    conn, virtual_event_id: int, challenge_id: int, limit: int
) -> list[dict]:
    return _all(
        conn,
        """
        WITH ranked AS (
            SELECT id, team_name, leader_name, leader_email,
                   total_score, deployed_link, github_repository_link,
                   export_created_at,
                   ROW_NUMBER() OVER (
                     ORDER BY total_score DESC NULLS LAST,
                              export_created_at ASC NULLS LAST,
                              id ASC
                   )::int AS rank
            FROM virtual_challenge_submission_rows
            WHERE event_id = :eid AND challenge_id = :cid
        )
        SELECT rank, team_name, leader_name, leader_email,
               total_score, deployed_link, github_repository_link,
               export_created_at AS submitted_at
        FROM ranked
        ORDER BY rank
        LIMIT :lim
        """,
        {"eid": int(virtual_event_id), "cid": int(challenge_id), "lim": int(limit)},
    )


def _virtual_cross_challenge_overlap(
    conn,
    virtual_event_id: int,
    challenge_1_id: int,
    challenge_2_id: int,
    top_n: int,
) -> dict[str, Any]:
    """Distinct C1 submitters; how many also submitted C2; same for top-N of C1 by best score."""
    eid = int(virtual_event_id)
    c1 = int(challenge_1_id)
    c2 = int(challenge_2_id)
    n = max(1, min(int(top_n), 100_000))
    row = conn.execute(
        text(
            """
            WITH c1 AS (
                SELECT DISTINCT leader_email_normalized AS em
                FROM virtual_challenge_submission_rows
                WHERE event_id = :eid AND challenge_id = :c1
            ),
            c2 AS (
                SELECT DISTINCT leader_email_normalized AS em
                FROM virtual_challenge_submission_rows
                WHERE event_id = :eid AND challenge_id = :c2
            ),
            best_c1 AS (
                SELECT leader_email_normalized AS em,
                       MAX(total_score) AS best_score,
                       MIN(export_created_at) AS first_at
                FROM virtual_challenge_submission_rows
                WHERE event_id = :eid AND challenge_id = :c1
                GROUP BY leader_email_normalized
            ),
            ranked AS (
                SELECT em, ROW_NUMBER() OVER (
                    ORDER BY best_score DESC NULLS LAST,
                             first_at ASC NULLS LAST,
                             em ASC
                ) AS rk
                FROM best_c1
            ),
            topn AS (SELECT em FROM ranked WHERE rk <= :n)
            SELECT
              (SELECT COUNT(*)::bigint FROM c1) AS c1_n,
              (SELECT COUNT(*)::bigint FROM c2) AS c2_n,
              (SELECT COUNT(*)::bigint FROM c1 INNER JOIN c2 USING (em)) AS c1_in_c2,
              (SELECT COUNT(*)::bigint FROM topn INNER JOIN c2 USING (em)) AS topn_in_c2
            """
        ),
        {"eid": eid, "c1": c1, "c2": c2, "n": n},
    ).mappings().fetchone() or {}
    return {
        "c1_distinct": int(row.get("c1_n") or 0),
        "c2_distinct": int(row.get("c2_n") or 0),
        "c1_also_in_c2": int(row.get("c1_in_c2") or 0),
        "top_n": n,
        "top_n_also_in_c2": int(row.get("topn_in_c2") or 0),
    }


def _in_person_to_virtual_overlap(
    conn, in_person_event_id: int, virtual_event_id: int
) -> dict[str, int]:
    row = conn.execute(
        text(
            f"""
            WITH ip AS (
                SELECT DISTINCT email_normalized AS em
                FROM {pw.TABLE_IN_PERSON_MDC}
                WHERE event_id = :ip_e
                  AND email_normalized IS NOT NULL AND btrim(email_normalized::text) <> ''
            ),
            v_reg AS (
                SELECT DISTINCT email_normalized AS em
                FROM {pw.TABLE_VIRTUAL_MDC}
                WHERE event_id = :v_e
                  AND email_normalized IS NOT NULL AND btrim(email_normalized::text) <> ''
            ),
            v_sub AS (
                SELECT DISTINCT leader_email_normalized AS em
                FROM virtual_challenge_submission_rows
                WHERE event_id = :v_e
                  AND leader_email_normalized IS NOT NULL
                  AND btrim(leader_email_normalized::text) <> ''
            )
            SELECT
              (SELECT COUNT(*)::bigint FROM ip) AS ip_n,
              (SELECT COUNT(*)::bigint FROM ip INNER JOIN v_reg USING (em)) AS ip_in_v_reg,
              (SELECT COUNT(*)::bigint FROM ip INNER JOIN v_sub USING (em)) AS ip_in_v_sub
            """
        ),
        {"ip_e": int(in_person_event_id), "v_e": int(virtual_event_id)},
    ).mappings().fetchone() or {}
    return {
        "in_person_distinct_emails": int(row.get("ip_n") or 0),
        "also_registered_virtual": int(row.get("ip_in_v_reg") or 0),
        "also_submitted_virtual": int(row.get("ip_in_v_sub") or 0),
    }


def _in_person_city_breakdown(conn, in_person_event_id: int) -> list[dict]:
    """Registrations per city. Prefers attendance_city, falls back to city for the label."""
    return _all(
        conn,
        f"""
        SELECT
          COALESCE(NULLIF(btrim(attendance_city), ''),
                   NULLIF(btrim(city), ''),
                   '<unknown>') AS city,
          COUNT(*)::bigint AS registrations,
          COUNT(*) FILTER (WHERE form_timestamp IS NOT NULL
                                 AND form_timestamp >= now() - interval '7 days')::bigint AS last_7_days
        FROM {pw.TABLE_IN_PERSON_MDC}
        WHERE event_id = :eid
        GROUP BY 1
        ORDER BY registrations DESC, city ASC
        """,
        {"eid": int(in_person_event_id)},
    )


def _in_person_summary(conn, in_person_event_id: int) -> dict[str, Any]:
    eid = int(in_person_event_id)
    total_reg = int(
        _scalar(conn, f"SELECT COUNT(*) FROM {pw.TABLE_IN_PERSON_MDC} WHERE event_id = :e", {"e": eid}) or 0
    )
    last7 = int(
        _scalar(
            conn,
            f"""
            SELECT COUNT(*) FROM {pw.TABLE_IN_PERSON_MDC}
            WHERE event_id = :e AND form_timestamp IS NOT NULL
              AND form_timestamp >= now() - interval '7 days'
            """,
            {"e": eid},
        )
        or 0
    )
    distinct_cities = int(
        _scalar(
            conn,
            f"""
            SELECT COUNT(DISTINCT COALESCE(NULLIF(btrim(attendance_city), ''),
                                            NULLIF(btrim(city), '')))
            FROM {pw.TABLE_IN_PERSON_MDC}
            WHERE event_id = :e
            """,
            {"e": eid},
        )
        or 0
    )
    main_row = conn.execute(
        text(
            """
            SELECT COUNT(*)::bigint AS total_subs,
                   COUNT(DISTINCT leader_email_normalized)::bigint AS unique_present,
                   COUNT(*) FILTER (WHERE deployed_link IS NOT NULL
                                          AND btrim(deployed_link) <> '')::bigint AS deployed,
                   COUNT(*) FILTER (WHERE total_score IS NOT NULL)::bigint AS scored,
                   COUNT(DISTINCT attendance_city_normalized)::bigint AS cities_with_subs
            FROM in_person_challenge_submission_rows
            WHERE event_id = :e AND sheet_kind = 'main'
            """
        ),
        {"e": eid},
    ).mappings().fetchone() or {}
    warmup_row = conn.execute(
        text(
            """
            SELECT COUNT(*)::bigint AS total,
                   COUNT(DISTINCT leader_email_normalized)::bigint AS uniq
            FROM in_person_challenge_submission_rows
            WHERE event_id = :e AND sheet_kind = 'warmup'
            """
        ),
        {"e": eid},
    ).mappings().fetchone() or {}
    return {
        "total_registrations": total_reg,
        "registrations_last_7_days": last7,
        "distinct_cities_in_registrations": distinct_cities,
        "main_total_submissions": int(main_row.get("total_subs") or 0),
        "people_present_distinct_emails_main": int(main_row.get("unique_present") or 0),
        "main_projects_deployed": int(main_row.get("deployed") or 0),
        "main_scored_count": int(main_row.get("scored") or 0),
        "main_cities_with_submissions": int(main_row.get("cities_with_subs") or 0),
        "warmup_total_submissions": int(warmup_row.get("total") or 0),
        "warmup_distinct_emails": int(warmup_row.get("uniq") or 0),
    }


def _in_person_top_main_teams(conn, in_person_event_id: int, limit: int) -> list[dict]:
    return _all(
        conn,
        """
        WITH ranked AS (
            SELECT id, team_name, leader_name, leader_email, attendance_city,
                   total_score, deployed_link, github_repository_link,
                   prompt_war_on, session_label, export_created_at,
                   ROW_NUMBER() OVER (
                     ORDER BY total_score DESC NULLS LAST,
                              export_created_at ASC NULLS LAST,
                              id ASC
                   )::int AS rank
            FROM in_person_challenge_submission_rows
            WHERE event_id = :eid AND sheet_kind = 'main'
        )
        SELECT rank, team_name, leader_name, leader_email, attendance_city,
               total_score, deployed_link, github_repository_link,
               prompt_war_on, session_label,
               export_created_at AS submitted_at
        FROM ranked
        ORDER BY rank
        LIMIT :lim
        """,
        {"eid": int(in_person_event_id), "lim": int(limit)},
    )


def main() -> None:
    p = argparse.ArgumentParser(description="Export PromptWars in-person + virtual stats to XLSX.")
    p.add_argument(
        "-o", "--output",
        type=Path, default=ROOT / "promptwars_stats.xlsx",
        help="Output .xlsx path (default: ./promptwars_stats.xlsx)",
    )
    p.add_argument("--in-person-event-id", type=int, default=pw.DEFAULT_IN_PERSON_EVENT_ID)
    p.add_argument("--virtual-event-id", type=int, default=pw.DEFAULT_VIRTUAL_EVENT_ID)
    p.add_argument(
        "--top-n", type=int, default=400,
        help="Top N rows per virtual challenge sheet, and in-person Top sheet uses 10 (default: 400).",
    )
    p.add_argument(
        "--virtual-c1-id", type=int, default=None,
        help="Override 'Challenge 1' id (default: auto-detected first challenge of the virtual event).",
    )
    p.add_argument(
        "--virtual-c2-id", type=int, default=None,
        help="Override 'Challenge 2' id (default: auto-detected second challenge of the virtual event).",
    )
    p.add_argument(
        "--virtual-c1-top-n", type=int, default=400,
        help="N for 'top N of Challenge 1 also submitted in Challenge 2' overlap (default: 400).",
    )
    args = p.parse_args()

    ip_e = int(args.in_person_event_id)
    v_e = int(args.virtual_event_id)
    top_n = max(1, int(args.top_n))
    overlap_n = max(1, int(args.virtual_c1_top_n))

    now_ist = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y-%m-%d %H:%M IST")

    with pw.engine.connect() as conn:
        ip_summary = _in_person_summary(conn, ip_e)
        ip_cities = _in_person_city_breakdown(conn, ip_e)
        ip_top10 = _in_person_top_main_teams(conn, ip_e, 10)
        ip_to_v = _in_person_to_virtual_overlap(conn, ip_e, v_e)
        v_total_reg = _virtual_event_total_registrations(conn, v_e)
        v_overall = _virtual_overall_submission_stats(conn, v_e)
        v_challenges = _virtual_challenges(conn, v_e)
        per_ch = _virtual_per_challenge_stats(conn, v_e, v_challenges)
        c1_id = args.virtual_c1_id if args.virtual_c1_id is not None else (v_challenges[0]["id"] if len(v_challenges) >= 1 else None)
        c2_id = args.virtual_c2_id if args.virtual_c2_id is not None else (v_challenges[1]["id"] if len(v_challenges) >= 2 else None)
        cross_ch: dict[str, Any] | None = None
        if c1_id is not None and c2_id is not None:
            cross_ch = _virtual_cross_challenge_overlap(conn, v_e, int(c1_id), int(c2_id), overlap_n)
        per_ch_top: dict[int, list[dict]] = {}
        for ch in v_challenges:
            cid = int(ch["id"])
            per_ch_top[cid] = _virtual_top_n_for_challenge(conn, v_e, cid, top_n)

    c1_meta = next((c for c in v_challenges if c1_id is not None and int(c["id"]) == int(c1_id)), None)
    c2_meta = next((c for c in v_challenges if c2_id is not None and int(c["id"]) == int(c2_id)), None)

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary")
    _kv_section(
        ws_summary,
        "PromptWars - snapshot",
        [
            ("Generated (IST)", now_ist),
            ("In-person event_id", ip_e),
            ("Virtual event_id", v_e),
        ],
    )
    _kv_section(
        ws_summary,
        "In-person",
        [
            ("Total registrations", ip_summary["total_registrations"]),
            ("New registrations (last 7 days)", ip_summary["registrations_last_7_days"]),
            ("Distinct cities (registrations)", ip_summary["distinct_cities_in_registrations"]),
            ("People present on event days (distinct emails on main sheets)",
                ip_summary["people_present_distinct_emails_main"]),
            ("Submissions (main challenge rows)", ip_summary["main_total_submissions"]),
            ("Projects deployed (main, deployed_link not blank)", ip_summary["main_projects_deployed"]),
            ("Cities with at least one main submission", ip_summary["main_cities_with_submissions"]),
            ("Warm-up submissions (info only)", ip_summary["warmup_total_submissions"]),
        ],
    )
    _kv_section(
        ws_summary,
        "Virtual",
        [
            ("Total registrations (now)", v_total_reg),
            ("Total submissions (all challenges)", v_overall["total_submissions"]),
            ("Unique submissions by leader email (all challenges)", v_overall["unique_submissions_by_email"]),
            ("Submissions with deployed link", v_overall["deployed_count"]),
            ("Submissions with a score", v_overall["scored_count"]),
            ("Number of challenges on this event", len(v_challenges)),
        ],
    )
    _kv_section(
        ws_summary,
        "In-person -> Virtual (same email on registration forms)",
        [
            ("In-person registered (distinct emails)", ip_to_v["in_person_distinct_emails"]),
            ("...also registered on Virtual", ip_to_v["also_registered_virtual"]),
            ("...also submitted in any Virtual challenge", ip_to_v["also_submitted_virtual"]),
        ],
    )
    if cross_ch is not None and c1_meta is not None and c2_meta is not None:
        _kv_section(
            ws_summary,
            "Virtual cross-challenge",
            [
                (f"Challenge 1: id={c1_meta['id']} title", c1_meta.get("title")),
                (f"Challenge 2: id={c2_meta['id']} title", c2_meta.get("title")),
                (f"Distinct submitters in Challenge {c1_meta['id']}", cross_ch["c1_distinct"]),
                (f"Distinct submitters in Challenge {c2_meta['id']}", cross_ch["c2_distinct"]),
                (
                    f"Of Challenge {c1_meta['id']} submitters, also submitted in Challenge {c2_meta['id']}",
                    cross_ch["c1_also_in_c2"],
                ),
                (
                    f"Of top {cross_ch['top_n']} of Challenge {c1_meta['id']} (by best score), also submitted in Challenge {c2_meta['id']}",
                    cross_ch["top_n_also_in_c2"],
                ),
            ],
        )
    elif len(v_challenges) < 2:
        _kv_section(
            ws_summary,
            "Virtual cross-challenge",
            [("Note", f"Only {len(v_challenges)} challenge(s) on this virtual event; need at least 2.")],
        )
    _autosize(ws_summary)

    ws_ip = wb.create_sheet("InPerson")
    _kv_section(
        ws_ip,
        "Registrations",
        [
            ("Total registrations", ip_summary["total_registrations"]),
            ("New registrations (last 7 days)", ip_summary["registrations_last_7_days"]),
            ("Distinct cities", ip_summary["distinct_cities_in_registrations"]),
        ],
    )
    _kv_section(
        ws_ip,
        "Event-day activity (Action Center main sheet imports)",
        [
            ("People present (distinct leader emails on main sheets)", ip_summary["people_present_distinct_emails_main"]),
            ("Submissions (rows on main sheets)", ip_summary["main_total_submissions"]),
            ("Projects deployed (deployed_link not blank)", ip_summary["main_projects_deployed"]),
            ("Submissions with a score", ip_summary["main_scored_count"]),
            ("Cities with at least one main submission", ip_summary["main_cities_with_submissions"]),
            ("Warm-up submissions (info only)", ip_summary["warmup_total_submissions"]),
            ("Warm-up distinct emails (info only)", ip_summary["warmup_distinct_emails"]),
        ],
    )
    _kv_section(
        ws_ip,
        "In-person -> Virtual (same email on registration forms)",
        [
            ("In-person registered (distinct emails)", ip_to_v["in_person_distinct_emails"]),
            ("...also registered on Virtual", ip_to_v["also_registered_virtual"]),
            ("...also submitted in any Virtual challenge", ip_to_v["also_submitted_virtual"]),
        ],
    )
    _autosize(ws_ip)

    ws_cities = wb.create_sheet("InPerson_Cities")
    _table(
        ws_cities,
        ["city", "registrations", "registrations_last_7_days"],
        [[c.get("city"), int(c.get("registrations") or 0), int(c.get("last_7_days") or 0)] for c in ip_cities],
        title="Registrations by city",
    )
    _autosize(ws_cities)

    ws_top = wb.create_sheet("InPerson_Top10")
    _table(
        ws_top,
        ["rank", "team_name", "leader_name", "leader_email", "attendance_city",
         "total_score", "deployed_link", "github_repository_link",
         "prompt_war_on", "session_label", "submitted_at"],
        [
            [
                int(r.get("rank") or 0),
                r.get("team_name"),
                r.get("leader_name"),
                r.get("leader_email"),
                r.get("attendance_city"),
                float(r["total_score"]) if r.get("total_score") is not None else None,
                r.get("deployed_link"),
                r.get("github_repository_link"),
                str(r["prompt_war_on"]) if r.get("prompt_war_on") is not None else None,
                r.get("session_label"),
                r["submitted_at"].isoformat() if r.get("submitted_at") and hasattr(r["submitted_at"], "isoformat") else r.get("submitted_at"),
            ]
            for r in ip_top10
        ],
        title="In-person main challenge - top 10 (top 3 highlighted)",
        highlight_top3=True,
        highlight_top10=True,
    )
    _autosize(ws_top)

    ws_v = wb.create_sheet("Virtual")
    _kv_section(
        ws_v,
        "Registrations & submissions (overall)",
        [
            ("Total registrations (now)", v_total_reg),
            ("Total submissions (all challenges)", v_overall["total_submissions"]),
            ("Unique submissions by leader email (all challenges)", v_overall["unique_submissions_by_email"]),
            ("Submissions with deployed link", v_overall["deployed_count"]),
            ("Submissions with a score", v_overall["scored_count"]),
            ("Number of challenges on this event", len(v_challenges)),
        ],
    )
    if cross_ch is not None and c1_meta is not None and c2_meta is not None:
        _kv_section(
            ws_v,
            "Cross-challenge",
            [
                (f"Challenge 1: id={c1_meta['id']} title", c1_meta.get("title")),
                (f"Challenge 2: id={c2_meta['id']} title", c2_meta.get("title")),
                (f"Distinct submitters in Challenge {c1_meta['id']}", cross_ch["c1_distinct"]),
                (f"Distinct submitters in Challenge {c2_meta['id']}", cross_ch["c2_distinct"]),
                (
                    f"Challenge {c1_meta['id']} submitters who also submitted in Challenge {c2_meta['id']}",
                    cross_ch["c1_also_in_c2"],
                ),
                (
                    f"Top {cross_ch['top_n']} of Challenge {c1_meta['id']} (by best score) who also submitted in Challenge {c2_meta['id']}",
                    cross_ch["top_n_also_in_c2"],
                ),
            ],
        )
    else:
        _kv_section(
            ws_v,
            "Cross-challenge",
            [("Note", "Need at least two challenges on this virtual event for cross-challenge stats.")],
        )
    _kv_section(
        ws_v,
        "In-person -> Virtual (same email on registration forms)",
        [
            ("In-person registered (distinct emails)", ip_to_v["in_person_distinct_emails"]),
            ("...also registered on Virtual", ip_to_v["also_registered_virtual"]),
            ("...also submitted in any Virtual challenge", ip_to_v["also_submitted_virtual"]),
        ],
    )
    _autosize(ws_v)

    ws_chs = wb.create_sheet("Virtual_Challenges")
    _table(
        ws_chs,
        [
            "challenge_id", "title", "status",
            "opens_at", "closes_at",
            "registrations_at_open", "registrations_at_close",
            "registrations_during_window",
            "total_submissions", "unique_submissions_by_email",
            "deployed_count", "scored_count",
        ],
        [
            [
                ch.get("challenge_id"),
                ch.get("title"),
                ch.get("status"),
                ch["opens_at"].isoformat() if ch.get("opens_at") and hasattr(ch["opens_at"], "isoformat") else ch.get("opens_at"),
                ch["closes_at"].isoformat() if ch.get("closes_at") and hasattr(ch["closes_at"], "isoformat") else ch.get("closes_at"),
                ch.get("registrations_at_open"),
                ch.get("registrations_at_close"),
                ch.get("registrations_during_window"),
                ch.get("total_submissions"),
                ch.get("unique_submissions_by_email"),
                ch.get("deployed_count"),
                ch.get("scored_count"),
            ]
            for ch in per_ch
        ],
        title="Per-challenge: registration snapshots & submission counts",
    )
    _autosize(ws_chs)

    used_names: set[str] = set(wb.sheetnames)
    for ch in v_challenges:
        cid = int(ch["id"])
        rows = per_ch_top.get(cid) or []
        if not rows:
            continue
        slug = _slug(ch.get("title"))
        candidate = _safe_sheet(f"V_C{cid}_{slug}")
        base = candidate
        i = 2
        while candidate in used_names:
            candidate = _safe_sheet(f"{base[:28]}_{i}")
            i += 1
        used_names.add(candidate)
        ws_ch = wb.create_sheet(candidate)
        _kv_section(
            ws_ch,
            "Challenge",
            [
                ("challenge_id", cid),
                ("title", ch.get("title")),
                ("status", ch.get("status")),
                ("opens_at", ch["opens_at"].isoformat() if ch.get("opens_at") and hasattr(ch["opens_at"], "isoformat") else ch.get("opens_at")),
                ("closes_at", ch["closes_at"].isoformat() if ch.get("closes_at") and hasattr(ch["closes_at"], "isoformat") else ch.get("closes_at")),
                ("rows shown (top N)", len(rows)),
            ],
        )
        _table(
            ws_ch,
            ["rank", "team_name", "leader_name", "leader_email",
             "total_score", "deployed_link", "github_repository_link", "submitted_at"],
            [
                [
                    int(r.get("rank") or 0),
                    r.get("team_name"),
                    r.get("leader_name"),
                    r.get("leader_email"),
                    float(r["total_score"]) if r.get("total_score") is not None else None,
                    r.get("deployed_link"),
                    r.get("github_repository_link"),
                    r["submitted_at"].isoformat() if r.get("submitted_at") and hasattr(r["submitted_at"], "isoformat") else r.get("submitted_at"),
                ]
                for r in rows
            ],
            title=f"Top {len(rows)} (top 10 highlighted, top 3 in gold)",
            highlight_top3=True,
            highlight_top10=True,
        )
        _autosize(ws_ch)

    out_path = args.output.resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
